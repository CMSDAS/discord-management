import sys
import os
import discord
import openpyxl
from fuzzywuzzy import process
from dotenv import load_dotenv
import gspread
load_dotenv()

class MyClient(discord.Client):
    async def on_ready(self):
        print('Logged on as {0}!'.format(self.user))
        server = client.guilds[0]
        wb = openpyxl.load_workbook('./data/Long_Exercise_Names_Short.xlsx')
        sheet_name = 'Inputs'
        sheet = wb[sheet_name]

        part_names = [v[0].value for v in sheet["B2:B100"]]
        # print(part_names)

        members = server.members
        print(dir(members[0]))
        cleaned_members = []
        for member in members:
            # print(member.name, member.nick, member.roles, member.id)
            isFacilitator = False
            for role in member.roles:
                if role.name == "facilitators" or role.name == "organisers":
                    isFacilitator = True
            if not isFacilitator:
                cleaned_members.append(member)
        member_names = [member.name for member in cleaned_members]
        print(member_names)

        # Open for writing
        wb = openpyxl.Workbook()
        dest_filename = 'discord_names.xlsx'
        ws1 = wb.active
        ws1.title = "Discord Names"
        ws1.cell(column=1, row=1, value='Name')
        ws1.cell(column=2, row=1, value='Match 1')
        ws1.cell(column=3, row=1, value='Score 1')
        ws1.cell(column=4, row=1, value='Match 2')
        ws1.cell(column=5, row=1, value='Score 2')
        ws1.cell(column=6, row=1, value='Match 3')
        ws1.cell(column=7, row=1, value='Score 3')

        row_number = 2
        for part_name in part_names:
            query = part_name
            choices = member_names
            matches = process.extract(query, choices, limit=3)
            ws1.cell(column=1, row=row_number, value=query)
            ws1.cell(column=2, row=row_number, value=matches[0][0])
            ws1.cell(column=3, row=row_number, value=matches[0][1])
            ws1.cell(column=4, row=row_number, value=matches[1][0])
            ws1.cell(column=5, row=row_number, value=matches[1][1])
            ws1.cell(column=6, row=row_number, value=matches[2][0])
            ws1.cell(column=7, row=row_number, value=matches[2][1])
            # print(f"{query}: Best among the above list: ",process.extractOne(query, choices))
            row_number += 1

        wb.save(filename = dest_filename)
        # print(server.roles)
        sys.exit()

client = MyClient()
client.run(os.getenv("DISCORD_TOKEN"))